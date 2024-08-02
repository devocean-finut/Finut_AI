import openai
import os
import json
import random
from openpyxl import load_workbook
from dotenv import load_dotenv

# .env 파일의 환경 변수 로드
load_dotenv()

# OpenAI API 키 설정
openai.api_key = os.getenv('OPENAI_API_KEY')

# economy_terms.json 파일에서 데이터를 읽어오기
with open('economy_terms.json', 'r', encoding='utf-8') as f:
    economy_terms = json.load(f)

def generate_quiz(term_info):
    term = term_info['term']
    description = term_info['description']

    messages = [
        {"role": "system", "content": "You are a helpful assistant that creates quiz questions."},
        {"role": "user",
         "content": f"Generate a true or false quiz question based on the following description.\nTerm: {term}\nDescription: {description}\nQuestion format: '[Description]', True or False?"}
    ]

    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=messages,
        max_tokens=100
    )

    quiz = response.choices[0].text.strip()
    return quiz

# 예시 퀴즈 생성 및 출력
if economy_terms:
    quiz = generate_quiz(random.choice(economy_terms))
    print(quiz)
else:
    print("Cannot generate quiz")

# 엑셀 파일 경로
# excel_path = '/Users/user/Desktop/Finut_Quiz.xlsx'

# 엑셀 파일 열기
# wb = load_workbook(excel_path)
# ws = wb.active

# 첫 번째 빈 행 찾기
# def get_first_empty_row(worksheet):
#     for row in range(2, worksheet.max_row + 2):
#         if worksheet.cell(row=row, column=1).value is None:
#             return row
#     return worksheet.max_row + 1

# 예시 퀴즈 생성 및 엑셀 파일에 저장
# if economy_terms:
#     for term_info in economy_terms:
#         term = term_info['term']
#         description = term_info['description']
#         question = generate_quiz(term, description)
#         answer = "T" if "True" in question else "F"
#
#         # 엑셀 파일에 데이터 쓰기
#         row = get_first_empty_row(ws)
#         ws[f'A{row}'] = term
#         ws[f'B{row}'] = description
#         ws[f'C{row}'] = question
#         ws[f'D{row}'] = answer
#
#     # 엑셀 파일 저장
#     wb.save(excel_path)
#     print("Quiz data successfully saved to the Excel file.")
# else:
#     print("No terms found. Please check the economy_terms.json file.")
